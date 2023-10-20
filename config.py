import pandas as pd
import re
from copy import copy
from number_to_text import num2text
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
def get_key_by_value(dictionary, value):
    keys = list(dictionary.keys())
    if value in dictionary.values():
        index = list(dictionary.values()).index(value)
        return keys[index]
    return None  # Если значение не найдено
def convert_numbers_to_words(sentence):
    sentence = str(sentence)
    sentence = sentence.replace('.',' . ')
    sentence = sentence.replace(',', ' . ')
    sentence = sentence.replace('.', 'точка')

    words = sentence.split()  # разделение предложения на отдельные слова
    converted_words = []  # список для хранения преобразованных слов

    for word in words:
        if word.isdigit():  # проверка, является ли слово числом
            converted_word = num2text(int(word))  # преобразование числа в слово
            converted_words.append(converted_word)
        else:
            converted_words.append(word)  # оставляем слово без изменений

    converted_sentence = ' '.join(converted_words)  # объединение слов обратно в предложение
    return converted_sentence

df = pd.read_excel('table.xlsx',decimal=',')

months=['январь','февраль','март','апрель','май','июнь','июль','август','сентябрь','октябрь','ноябрь','декабрь']
items=[]
sootvetstvie={}
dwords = ['пао','оао','ао','ооо']
for item in pd.array(df['Заказчик']):
    converted = convert_numbers_to_words(item)
    converted = converted.replace('"','')
    converted = converted.replace('-', ' ')
    converted = converted.replace(':', ' ')
    converted=converted.lower()
    converted = converted.split()
    words = [x for x in converted if x not in dwords]
    converted=' '.join(words)
    items.append(converted)
    sootvetstvie[item] = converted
print(sootvetstvie)
VA_ALIAS = []

VA_TBR = ('для')


VA_CMDS = {
    'show1':('выполнение договоров'),
    'show2':('отгрузки товаров'),
    'comment':('добавь комментарий'),
    'sendmail':('отправь сообщение'),
    'stop':('стоп'),
    'stop2':('отмена')
}

from openpyxl.utils.dataframe import dataframe_to_rows
def savetable(input,output,df):
    source_book = openpyxl.load_workbook(input)
    # Загрузка исходного листа
    source_sheet = source_book['Sheet']
    # Создание новой книги
    target_book = openpyxl.Workbook()
    # Создание нового листа
    target_sheet = target_book.active
    ws=target_book.active
    for c in ['E','G','H','I']:
        ws.column_dimensions[c].width = 20
    # Перенос стилей ячеек
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            target_cell.font = copy(cell.font)
            target_cell.border = copy(cell.border)
            target_cell.fill = copy(cell.fill)
            target_cell.number_format = copy(cell.number_format)
            target_cell.protection = copy(cell.protection)
            target_cell.alignment = copy(cell.alignment)

    # Сохранение изменений в целевой книге
    target_book.save(output)
#
import configparser
cfg = configparser.ConfigParser()
cfg.read('config.ini')
sections = cfg.sections()
for option in cfg.options('NAMES'):
    VA_ALIAS.append(cfg.get('NAMES', option))
VA_ALIAS = tuple(VA_ALIAS)
departments={}
for item in sections:
    converted = convert_numbers_to_words(item)
    converted = converted.replace('"','')
    converted = converted.replace('-', ' ')
    converted = converted.replace(':', ' ')
    converted=converted.lower()
    departments[item] = converted
